import pandas as pd
from openpyxl import load_workbook
from datetime import datetime



def read_replacements(wb):
    sheet_replacements = wb["RULES"]  # Replace with the actual worksheet name

# Create an empty dictionary to store the data
    replacements = []

# Iterate through the rows in the worksheet
    for row in sheet_replacements.iter_rows(values_only=True):
        values = [row[i] for i in range(0, len(row))]
        replacements.append(values)
  
    return replacements

def main():
    
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S")

    file_path = r"C:\Users\rcxsm\Documents\xls\masterfinance_2023.xlsx"  # Replace with your file path
    wb = load_workbook(file_path)
    sheet = wb["INVOER"]

    # make a backup

    output_file_path = f'backup_{formatted_datetime}.xlsx'  # Replace with your desired output file path
    wb.save(output_file_path)
    print (f"Backup saved as {output_file_path}")
    replacements = read_replacements(wb)
   
    # Insert a new column with row numbers
    sheet.insert_cols(1)
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        sheet.cell(row=row_number + 1, column=1, value=row_number+1)

    # Loop through the rows in the worksheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        # Apply the replacements
        for r in replacements:
    
            try:
                if r[5] is None:
                    pass

                    # row[12] = income_expenses -> to see if there is already a category
                    if r[1].lower() in str(row[11]).lower() and row[12] is None:
                        print(f"{row[11]} x {row[8]}")
                        sheet.cell(row=row[0], column=13, value=r[2])
                        sheet.cell(row=row[0], column=14, value=r[3])
                        sheet.cell(row=row[0], column=15, value=r[4])
                else:
                    if r[1].lower() in str(row[11]).lower() and row[12] is None and row[8] == r[5]:
                        print(f"{row[11]} x {row[8]}")
                        sheet.cell(row=row[0], column=13, value=r[2])
                        sheet.cell(row=row[0], column=14, value=r[3])
                        sheet.cell(row=row[0], column=15, value=r[4])
            except:
                # Probably no description for this row
                pass

    sheet.delete_cols(1)

    wb.save(file_path)

if __name__ == "__main__":
    main()